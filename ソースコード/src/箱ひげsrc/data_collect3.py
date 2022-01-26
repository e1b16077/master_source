##########################################################
#  課題ごと ＆ work81とsample81を区別しているやつ#(保留中)（間違っているので修正する必要あり）

####################################

# 学生ごとの修正ログ数
# データベースの読み込み
import pandas as pd
import openpyxl
from datetime import datetime, time
import re
print("どの学科にしますか")
a = input()


book = openpyxl.load_workbook(
    '../データベース作成（出席多いと削除）/' + a + '_pure.xlsx', read_only=True)


# 各授業回で作成すべき課題名が格納されているファイルを読み込む
book2 = openpyxl.load_workbook('java2020_analysis.xlsx', read_only=True)
name2 = book2.sheetnames
# print(name2)
sheet2 = book2[name2[0]]
data2 = sheet2.values
data2 = list(data2)
data_df2 = pd.DataFrame(data2[1:], columns=data2[0])


# シートを取得
name = book.sheetnames
# log_sum = []
user_name = []
lec_sum = []


for j in range(len(name)):
    try:
        sheet = book[name[j+1]]
        data = sheet.values
        data = list(data)
        data_df = pd.DataFrame(data[1:], columns=data[0])
        user_name.append(name[j+1])
        count = 0

        # kadai名から_より前のものの列を追加する
        data_df['kadai_a'] = data_df['kadai'].apply(
            lambda x: re.sub(r'_.*', '', x))

        # 大まかな流れとしては、kadai_aでグループ化して、それの長さが1以下の場合削除
        s = len(data_df)
        # data_df = data_df.reset_index()
        count = []
        o = 0
        for i in range(s):
            lecture_name = data_df.iat[i, 3]
            kadai_name = data_df.iat[i, 11]
            # print(kadai_name)
            # lecture_nameが一致する行を抜き出す
            data_df3 = data_df2[data_df2["lec"] == lecture_name]
            if o == 0:
                co_data_df4 = data_df3

            if len(co_data_df4[co_data_df4['Workfile1'].str.contains(kadai_name, na=False)]) == 1:
                co_data_df4 = co_data_df4[~co_data_df4['Workfile1'].str.contains(
                    kadai_name, na=False)]

            elif len(co_data_df4[co_data_df4['Workfile2'].str.contains(kadai_name, na=False)]) == 1:
                co_data_df4 = co_data_df4[~co_data_df4['Workfile2'].str.contains(
                    kadai_name, na=False)]

            elif len(co_data_df4[co_data_df4['Workfile3'].str.contains(kadai_name, na=False)]) == 1:
                co_data_df4 = co_data_df4[~co_data_df4['Workfile3'].str.contains(
                    kadai_name, na=False)]

            elif len(co_data_df4[co_data_df4['Workfile4'].str.contains(kadai_name, na=False)]) == 1:
                co_data_df4 = co_data_df4[~co_data_df4['Workfile4'].str.contains(
                    kadai_name, na=False)]

            elif len(co_data_df4[co_data_df4['Workfile5'].str.contains(kadai_name, na=False)]) == 1:
                co_data_df4 = co_data_df4[~co_data_df4['Workfile5'].str.contains(
                    kadai_name, na=False)]

            else:
                o += 1
                continue
            o += 1

        lec_sum.append(len(data_df2) - len(co_data_df4))
        # if data_df3[~data_df3['Workfile1'].str.contains(kadai_name, na=False)]:

        # if len(data_df3[data_df3['Workfile1'].str.contains(kadai_name, na=False)]) == 0 and len(data_df3[data_df3['Workfile2'].str.contains(kadai_name, na=False)]) == 0 and len(data_df3[data_df3['Workfile3'].str.contains(kadai_name, na=False)]) == 0 and len(data_df3[data_df3['Workfile4'].str.contains(kadai_name, na=False)]) == 0 and len(data_df3[data_df3['Workfile5'].str.contains(kadai_name, na=False)]) == 0:
        #     count.append(i)

    except IndexError:
        break
print(lec_sum)

# log_sum = pd.DataFrame({a + '_sum_log': log_sum})
lec_sum = pd.DataFrame({a + '_lec_sum': lec_sum})
user_name = pd.DataFrame({a + '_use_name': user_name})
lec_sum = user_name.join(lec_sum)

# print(li[1][0:2])
# excelへ出力

# if len(lec_sum) > 0:
#     with pd.ExcelWriter('data_pure.xlsx', engine="openpyxl", mode="a") as writer:
#         lec_sum.to_excel(writer, sheet_name=a)
