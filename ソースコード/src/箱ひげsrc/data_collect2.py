##########################################################
#  課題ごと #
####################################


# 学生ごとの修正ログ数
# データベースの読み込み
import pandas as pd
import openpyxl
from datetime import datetime, time
import re
print("どの学科にしますか")
a = input()
# 読み込んだexcelのシート名取得
# book = openpyxl.load_workbook(
#     '../データベース作成（出席多いと削除）/01' + a + '.xlsx', read_only=True)
# book = openpyxl.load_workbook(
#     '../データベース作成（出席多いと削除）' + a + '_saisyu.xlsx', read_only=True)
# book = openpyxl.load_workbook(
#     '../データベース作成（出席多いと削除）/' + a + '_saisyu.xlsx', read_only=True)

# book = openpyxl.load_workbook(
#     '../データベース作成（出席多いと削除）/' + a + '_after_review.xlsx', read_only=True)


book = openpyxl.load_workbook(
    '../データベース作成（出席多いと削除）/' + a + '_pure+.xlsx', read_only=True)

# シートを取得
name = book.sheetnames
# log_sum = []
user_name = []
lec_sum = []


for j in range(len(name)):
    try:
        sheet = book[name[j+1]]
        data = sheet.values
        data = list(data)
        data_df = pd.DataFrame(data[1:], columns=data[0])
        user_name.append(name[j+1])
        count = 0

        # kadai名から_より前のものの列を追加する
        data_df['kadai_a'] = data_df['kadai'].apply(
            lambda x: re.sub(r'_.*', '', x))

        # 全てを大文字化（後から削除可能）
        data_df['kadai_a'] = data_df['kadai_a'].str.upper()

        # print(data_df['kadai_a'])

        data_df.drop_duplicates(subset=['kadai_a'], inplace=True)

        lec_sum.append(len(data_df))
    except IndexError:
        break


# log_sum = pd.DataFrame({a + '_sum_log': log_sum})
lec_sum = pd.DataFrame({a + '_lec_sum': lec_sum})
user_name = pd.DataFrame({a + '_use_name': user_name})
lec_sum = user_name.join(lec_sum)

# print(li[1][0:2])
# excelへ出力
if len(lec_sum) > 0:
    with pd.ExcelWriter('data_pure+.xlsx', engine="openpyxl", mode="a") as writer:
        lec_sum.to_excel(writer, sheet_name=a)
