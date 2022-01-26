# 一度でもレビューを残した人

# 学生ごとの修正ログ数
# データベースの読み込み
import pandas as pd
import openpyxl
from datetime import datetime, time
import re
import math
import matplotlib.pyplot as plt
# print("どの学科にしますか")
# a = input()
# 読み込んだexcelのシート名取得
book = openpyxl.load_workbook('nyuryoku.xlsx', read_only=True)

# シートを取得
name = book.sheetnames
# log_sum = []
user_name = []
lec_sum = []
count = 0

# 対応表
book2 = openpyxl.load_workbook('taiouhyou.xlsx', read_only=True)
name2 = book2.sheetnames
sheet2 = book2[name2[0]]
data2 = sheet2.values
data2 = list(data2)
data_df2 = pd.DataFrame(data2[1:], columns=data2[0])

# 成績読み込み
book3 = openpyxl.load_workbook('analysis.xlsx', read_only=True)
name3 = book3.sheetnames
sheet3 = book3[name3[0]]
data3 = sheet3.values
data3 = list(data3)
data_df3 = pd.DataFrame(data3[1:], columns=data3[0])
print(len(data_df3))

# 出席10回未満読み込み
book4 = openpyxl.load_workbook('kesseki.xlsx', read_only=True)
name4 = book4.sheetnames
sheet4 = book4[name4[0]]
data4 = sheet4.values
data4 = list(data4)
data_df4 = pd.DataFrame(data4[1:], columns=data4[0])
kessekisya = data_df4.values.tolist()
kessekisya = sum(kessekisya, [])
print(len(kessekisya))


# data_df3からkessekisyaに一致しないものを取り出す
data_df3 = data_df3[~data_df3['hsid'].isin(kessekisya)]
print(len(data_df3))

# dataframeを辞書型に変更する
result_dict = dict(data_df2[["ZID", "hsid"]].values)
# print(result_dict)
# print(data_df2)
for j in range(len(name)):
    try:

        sheet = book[name[j]]
        data = sheet.values
        data = list(data)
        data_df = pd.DataFrame(data[1:], columns=data[0])
        user_name.append(name[j])
        # print(name[j])

        data_df["レビュアー"] = data_df['レビュアー'].map(result_dict)

        if count == 0:
            list_sample = data_df['レビュアー'].to_list()
        else:
            # new_df = pd.concat([new_df, data_df], axis=0)
            list_sample = list_sample + data_df['レビュアー'].to_list()

        count += 1
        # print(data_df[data_df['label'] == 1])

    except IndexError:
        break

# 重複を削除
purpo1 = list(set(list_sample))


# 頭文字がeから始まらないものを削除する処理
newlist1 = [x for x in purpo1 if str(x) != 'nan' and str(x) != 'None']


# kessekiに含まれるものをlistから削除する処理を追加する
for w in kessekisya:
    try:
        # print(w)
        newlist1.remove(w)
    except ValueError:
        pass


s1 = data_df3[data_df3['hsid'].isin(newlist1)]

print(s1["SAチェック1"].to_list())
print("次")
print(s1["SAチェック2"].to_list())
print("次")
print(s1["SAチェック3"].to_list())
data = (s1["SAチェック1"].to_list(), s1["SAチェック2"].to_list(), s1["SAチェック3"].to_list())

plt.rcParams['font.size'] = 15
fig1, ax1 = plt.subplots()
# ax1.set_title('学科ごとのログ数')
ax1.set_xticklabels(['レポート1', 'レポート2',
                     'レポート3'], fontname="MS Gothic")

# plt.title("授業に10回以上参加し1つの課題に対して2回以上編集している学生のレビュー後課題編集数の分布",
#           y=-0.23, fontname="MS Gothic")
# plt.title("クラス毎のレビュー後課題実施数",
#           y=-0.23, fontname="MS Gothic")
plt.subplots_adjust(bottom=0.2, top=0.95)

ax1.boxplot(data, sym="")
plt.savefig("300_dpi_scatter.png", format="png", dpi=300)
plt.show()
