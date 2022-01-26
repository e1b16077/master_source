##########################################################
# ソースコードの差分を表示するプログラム
####################################

# データベースの読み込み
import pandas as pd
import openpyxl
from datetime import datetime, time
import re
import difflib


print("どの学科にしますか：IS科は1,IC科は2,IN科は3")
a = input()

if a == "1":
    a = "IS"
elif a == "2":
    a = "IC"
elif a == "3":
    a = "IN"
else:
    pass

print("ユーザ名の入力")
b = input()
# 読み込んだexcelのシート名取得
# レビュー後のシートの処理
book1 = openpyxl.load_workbook(
    a + '_pure+.xlsx', read_only=True)
# シートを取得
name1 = book1.sheetnames
sheet = book1[b]
data1 = sheet.values
data1 = list(data1)
data_df1 = pd.DataFrame(data1[1:], columns=data1[0])

# コミット確認
# レビュー前のシートの処理
book2 = openpyxl.load_workbook(
    a + '_test_last.xlsx', read_only=True)
# シートを取得
name2 = book2.sheetnames
sheet = book2[b]
data2 = sheet.values
data2 = list(data2)
data_df2 = pd.DataFrame(data2[1:], columns=data2[0])
# print(data_df2["kadai"])

# ここまで前置き

data_df1['kadai_a'] = data_df1['kadai'].apply(
    lambda x: re.sub(r'_.*', '', x))

data_df2['kadai_a'] = data_df2['kadai'].apply(
    lambda x: re.sub(r'_.*', '', x))

for i in range(100):
    print(data_df1['lecname'].unique())
    print("対象lecは？")
    lec = input()
    lec = "lec" + lec

    df3 = data_df1.query(
        'lecname == @lec')
    # print(df3['kadai_a'].unique())
    df_series = df3['kadai_a'].unique()
    df_series = df_series.tolist()

    # 辞書にするためのリストを用意する
    length = len(df_series)
    soeji_hairetu = []
    for soeji in range(1, length+1):
        soeji_hairetu.append(soeji)

    kadai_itiran = dict(zip(soeji_hairetu, df_series))
    print(kadai_itiran)
    print("対象課題は？")
    kadai = input()

    kadai = kadai_itiran[int(kadai)]

    while True:
        # 対象データの抽出が完了
        obst1 = data_df1.query(
            'lecname == @lec and kadai_a == @kadai')

        obst2 = data_df2.query(
            'lecname == @lec and kadai_a == @kadai')

        if len(obst2) == 0:
            print("この学生は、この課題を授業中に取り組んでいません")

        final = pd.concat([obst2, obst1])
        final = final.sort_values('time')

        # ここを追加
        final.reset_index(inplace=True, drop=True)
        final.index = final.index + 1
        print(final)
        print("どことの差分を求めますか?1から" + str(len(final)) + "です(授業回選択に戻るには「9 9」と入力)")

        num1, num2 = (int(x) for x in input().split())

        if num1 == 9 and num2 == 9:
            break

        s1 = str(final.iat[num1-1, 5])
        s2 = str(final.iat[num2-1, 5])
        output1 = s1.split("\n")
        output2 = s2.split("\n")

        d = difflib.Differ()
        diff = d.compare(output1, output2)
        print('\n'.join(diff))
        stop = input()
