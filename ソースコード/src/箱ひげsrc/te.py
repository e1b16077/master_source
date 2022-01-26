# 学生ごとの修正ログ数
# データベースの読み込み
import pandas as pd
import openpyxl
from datetime import datetime, time
import matplotlib.pyplot as plt
# 読み込んだexcelのシート名取得
# book = openpyxl.load_workbook('after_review_data.xlsx', read_only=True)
book = openpyxl.load_workbook('IS+IC.xlsx', read_only=True)
# シートを取得
name = book.sheetnames
log_sum = []
user_name = []
graph = []
s = ["IS", "IN"]
for j in range(len(name)):
    try:
        sheet = book[name[j+1]]
        data = sheet.values
        data = list(data)
        data_df = pd.DataFrame(data[1:], columns=data[0])

        graph.append(data_df[[s[j]+"_use_name", s[j]+"_lec_sum"]])

    except IndexError:
        break
# print(len(graph[2]))



# g = graph[2][["IC_sum_log", "IS_sum_log", "IN_sum_log"]]

g = graph[0]["IS_lec_sum"].tolist()
r = graph[1]["IN_lec_sum"].tolist()


data = (g, r)

plt.rcParams['font.size'] = 15
fig1, ax1 = plt.subplots()
# ax1.set_title('学科ごとのログ数')
ax1.set_xticklabels(['レビュー実施クラス',
                     'レビュー未実施クラスC'], fontname="MS Gothic")

# plt.title("授業に10回以上参加し1つの課題に対して2回以上編集している学生のレビュー後課題編集数の分布",
#           y=-0.23, fontname="MS Gothic")
# plt.title("クラス毎のレビュー後課題実施数",
#           y=-0.23, fontname="MS Gothic")
# plt.subplots_adjust(bottom=0.2, top=0.95)

ax1.boxplot(data, sym="")
plt.savefig("300_dpi_scatter.png", format="png", dpi=300)
plt.show()
