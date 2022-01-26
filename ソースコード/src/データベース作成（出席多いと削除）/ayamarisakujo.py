##########################################################
# 編集文字数一文字以下を削除したデータベースの中で1つの課題に対して編集が一度しか行われなかったものを削除する
####################################

# branchテスト
# 学生ごとの修正ログ数
# データベースの読み込み
import pandas as pd
import openpyxl
from datetime import datetime, time
import re
print("どの学科にしますか")
a = input()
# 読み込んだexcelのシート名取得
book = openpyxl.load_workbook(
    a + '_saisyu.xlsx', read_only=True)
# シートを取得
name = book.sheetnames


user_name = []
lec_sum = []

# 各授業回で作成すべき課題名が格納されているファイルを読み込む
book2 = openpyxl.load_workbook('java2020_analysis.xlsx', read_only=True)
name2 = book2.sheetnames
# print(name2)
sheet2 = book2[name2[0]]
data2 = sheet2.values
data2 = list(data2)
data_df2 = pd.DataFrame(data2[1:], columns=data2[0])

# print(data_df2)

with pd.ExcelWriter('IC_pure+.xlsx', engine="openpyxl", mode="a") as writer:
    for j in range(len(name)):
        try:
            sheet = book[name[j+1]]
            data = sheet.values
            data = list(data)
            data_df = pd.DataFrame(data[1:], columns=data[0])
            user_name.append(name[j+1])
            count = 0

            # kadai名から_より前のものの列を追加する
            data_df['kadai_a'] = data_df['kadai'].apply(
                lambda x: re.sub(r'_.*', '', x))

            # 大まかな流れとしては、kadai_aでグループ化して、それの長さが1以下の場合削除
            s = len(data_df)
            # data_df = data_df.reset_index()
            count = []
            for i in range(s):
                lecture_name = data_df.iat[i, 3]
                kadai_name = data_df.iat[i, 11]
                print(kadai_name)
                # lecture_nameが一致する行を抜き出す
                data_df3 = data_df2[data_df2["lec"] == lecture_name]
                data_df3 = data_df3.fillna("kessontidesu")
                kadai_name = kadai_name + ".java"

                # 部分一致だとファイル名がaとかでも一致していると判断されたため完全一致に変更した
                if len(data_df3[data_df3['Workfile1'] == kadai_name]) == 0 and len(data_df3[data_df3['Workfile2'] == kadai_name]) == 0 and len(data_df3[data_df3['Workfile3'] == kadai_name]) == 0 and len(data_df3[data_df3['Workfile4'] == kadai_name]) == 0 and len(data_df3[data_df3['Workfile5'] == kadai_name]) == 0:
                    count.append(i)

            copy_df = data_df.drop(count)

            copy_df = copy_df[["SID", "time", "lecname", "kadai", "contents",
                               "sumletter", "diff_sumletter", "loc", "diff_loc", "再編集時期"]]

            if len(copy_df) > 0:
                copy_df.to_excel(writer, sheet_name=name[j+1])

        except IndexError:
            break
